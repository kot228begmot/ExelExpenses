import openpyxl
import progressbar
class ExelExpenses:
    def __init__(self):
        self. _book = openpyxl.open("C:/Users/apec9/Dropbox/ExelExpenses/test.xlsx", read_only=True)
        self.list_number = self._book.worksheets[0]
        self._storage_of_category = {}
    def expenses_processing(self, month, year):
        error_positions = []
        start_end_month_line = self.find_border_of_the_month(month, year)
        self.category_counter(start_end_month_line)
        a=1

    def find_border_of_the_month(self, month, year):
        _position_of_start_line = 2
        _position_of_end_line = self.list_number.max_row
        _position_of_general_data_column = 2
        flag_find_year = False
        start_year_line = 0
        start_month_line = 0

        if self.list_number[_position_of_start_line][_position_of_general_data_column].value.year == year:
            start_year_line = _position_of_start_line
            flag_find_year = True
        else:
            while flag_find_year == False:
                if self.list_number[_position_of_start_line][
                    _position_of_general_data_column].value.year == year - 1 and \
                        self.list_number[_position_of_start_line + 1][
                            _position_of_general_data_column].value.year == year:
                    start_year_line = _position_of_start_line + 1
                    flag_find_year = True
                    break
                middle = (_position_of_start_line + _position_of_end_line) // 2
                print(middle)
                if self.list_number[middle][_position_of_general_data_column].value.year < year:
                    _position_of_start_line = middle
                else:
                    _position_of_end_line = middle
        flag_find_start_month = False
        _position_of_start_line = start_year_line
        _position_of_end_line = self.list_number.max_row
        if month == 1:
            start_month_line = start_year_line
            flag_find_start_month = True
        else:
            while flag_find_start_month == False:
                if (self.list_number[_position_of_start_line][
                        _position_of_general_data_column].value.month == month - 1 and \
                    self.list_number[_position_of_start_line + 1][
                        _position_of_general_data_column].value.month == month) or (
                        middle == start_year_line and self.list_number[_position_of_start_line - 1][
                    _position_of_general_data_column].value.month != month):
                    start_month_line = _position_of_start_line
                    flag_find_start_month = True
                    break
                middle = (_position_of_start_line + _position_of_end_line) // 2
                print(middle)
                if self.list_number[middle][_position_of_general_data_column].value.month < month:
                    _position_of_start_line = middle
                else:
                    _position_of_end_line = middle

        flag_find_end_month = False
        _position_of_end_line = self.list_number.max_row
        _position_of_start_line = start_month_line

        while flag_find_end_month == False:
            if (self.list_number[middle][
                _position_of_general_data_column].value.month == month and \
                    self.list_number[middle + 1][
                        _position_of_general_data_column].value.month != month):
                return [start_year_line, middle]
            middle = (_position_of_start_line + _position_of_end_line) // 2
            print(middle)
            if self.list_number[middle][_position_of_general_data_column].value.month > month:
                _position_of_end_line = middle
            else:
                _position_of_start_line = middle


    def category_counter(self, start_end_line):
        pb = progressbar.progressbar(range(start_end_line[0], start_end_line[1] + 1))
        _column_category = 4
        _column_under_category = 5
        _column_category_value = 6
        for row in range(start_end_line[0], start_end_line[1] + 1):
            next(pb)
            category = self.list_number[row][_column_category].value.lower()
            if category not in self._storage_of_category:
                self._storage_of_category.setdefault(category, {})
            if self.list_number[row][_column_under_category].value == None and len(self._storage_of_category[category]) == 0:
                self._storage_of_category[category].setdefault('noncategory', 0)
            if self.list_number[row][_column_under_category].value == None and len(self._storage_of_category[category]) != 0:
                self._storage_of_category[category].setdefault('noncategory', self.list_number[row][_column_category_value].value)
                self._storage_of_category[category]['noncategory'] =  \
                    self._storage_of_category[category]['noncategory'] + self.list_number[row][_column_category_value].value
            elif self.list_number[row][_column_under_category].value.lower() not in self._storage_of_category[category]:
                self._storage_of_category[category].setdefault(self.list_number[row][_column_under_category].value.lower(),self.list_number[row][_column_category_value].value)
            elif self.list_number[row][_column_under_category].value.lower()  in self._storage_of_category[category]:
                self._storage_of_category[category][self.list_number[row][_column_under_category].value.lower()] =  \
                    self._storage_of_category[category][self.list_number[row][_column_under_category].value.lower()]\
                                                                          + self.list_number[row][_column_category_value].value
        summa = []
        all_category = list(self._storage_of_category.keys())
        for i in range(0, len(self._storage_of_category)):
            all_category_ = list(self._storage_of_category[all_category[i]].keys())
            for j in all_category_:
                summa.append(self._storage_of_category[all_category[i]][j])
        print('Контрольная сумма:', sum(summa))
        a=1


xui = ExelExpenses()
a = xui.expenses_processing(11, 2022)



